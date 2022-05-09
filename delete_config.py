from zeep import Client
from zeep.cache import SqliteCache
from zeep.transports import Transport
from zeep.exceptions import Fault
from zeep.plugins import HistoryPlugin
from requests import Session
from requests.auth import HTTPBasicAuth
from urllib3 import disable_warnings
from urllib3.exceptions import InsecureRequestWarning
from lxml import etree
import sys
import boto3
from pathlib import Path
import openpyxl
import requests
import json
from credentials import credentials
from messages import messages, messageText, webexHeaders
from get_tokens import get_my_access_token
from delete_ldap import delete_ldap_user

#this script is run any time the delete command is sent to the bot
def delete_inbox (id_of_user, client):
    response = client.deregister_from_work_mail(
        OrganizationId=credentials["aws_organization_id"],
        EntityId=id_of_user,
    )
    print('deregister response: ', response)
    response = client.delete_user(
        OrganizationId=credentials["aws_organization_id"],
        UserId=id_of_user
    )
    print('delete response: ', response)

def deleteInbox(userid):
   #delete Workmail inboxes
   found = False
   print('analyzing ', userid)
   next_token = 'null'
   # initialize boto3 client
   boto3_client = boto3.client('workmail', region_name=credentials["aws_region_name"], aws_access_key_id=credentials["aws_access_key_id"],
                               aws_secret_access_key=credentials["aws_secret_access_key"])

   boto3_response = boto3_client.list_users(OrganizationId=credentials["aws_organization_id"])
   r = boto3_response['Users']
   h = len(r)
   for j in range(h):
      if r[j]['Name'] == userid and r[j]['State'] == 'ENABLED':
         found = True
         id_of_user = r[j]['Id']
         delete_inbox(id_of_user, boto3_client)
         break
   page = 1
   while found == False:
      if 'NextToken' in boto3_response:
          page = page + 1
          next_token = boto3_response['NextToken']
          boto3_response = boto3_client.list_users(OrganizationId=credentials["aws_organization_id"], NextToken=next_token)
          r = boto3_response['Users']
          h = len(r)
          for j in range(h):
            if r[j]['Name'] == userid and r[j]['State'] == 'ENABLED':
                found = True
                print('Found in page: ', page)
                id_of_user = r[j]['Id']
                delete_inbox(id_of_user, boto3_client)
                break
      else:
          print('Not found')
          break



def deleteLicenses(email, roomId):

    bearer = get_my_access_token(roomId)
    headers = {'Authorization': 'Bearer ' + bearer}
    response = requests.get(license_url, headers=headers)
    response = response.json()
    licenses_list = response['items']
    l = len(licenses_list)
    for i in range(l):
        if licenses_list[i]['name'] == "Webex Calling - Professional":
            wxc_license = licenses_list[i]['id']
            break
    response = requests.get(user_details_url + email, headers=headers)
    response = response.json()
    print('response for email: ', email, '\n', response)
    agent_id = response['items'][0]['id']
    agent_licenses = response['items'][0]['licenses']
    if wxc_license in agent_licenses:
        agent_licenses.remove(wxc_license)
        data_from_people_query = response['items'][0]
        if 'displayName' not in data_from_people_query:
            data_from_people_query['displayName'] = ''
        data_from_people_query['licenses'] = agent_licenses
        content_type_headers = headers
        content_type_headers['Content-Type'] = 'application/json'
        payload = json.dumps(data_from_people_query)
        print('payload is: ', payload)
        url = people_url + data_from_people_query['id'] + '?callingData=true'
        print('sending request to: ', url)
        response = requests.put(people_url + data_from_people_query['id'] + '?callingData=true', data=payload,
                                headers=content_type_headers)
        print(response)
        print('put response is: ', response.json())


def deleteDI (userid,tel,partition):
      #Delete lines, users and soft clients in UCM
      phone_prefixes = ['CSF', 'TAB', 'TCT', 'BOT']
      phones = {'CSF': 'Cisco Unified Client Services Framework', 'TAB': 'Cisco Jabber for Tablet',
                'TCT': 'Cisco Dual Mode for iPhone', 'BOT': 'Cisco Dual Mode for Android'}
      wsdl = credentials["wsdl"]
      location = 'https://{host}:8443/axl/'.format(host=host)
      binding = "{http://www.cisco.com/AXLAPIService/}AXLAPIBinding"
      session = Session()
      session.verify = False
      session.auth = HTTPBasicAuth(username, password)

      transport = Transport(cache=SqliteCache(), session=session, timeout=20)
      history = HistoryPlugin()
      client = Client(wsdl=wsdl, transport=transport, plugins=[history])
      service = client.create_service(binding, location)


      def show_history():
          for item in [history.last_sent, history.last_received]:
              print(etree.tostring(item["envelope"], encoding="unicode", pretty_print=True))


      line_data = {
         'pattern' : tel,
         'routePartitionName': partition
               }
      try:
          resp = service.removeLine(**line_data)
          print('\nremoveLine response:')
          print(resp, '\n')
      except Fault as err:
          print( f'Zeep error: removeLine: { err }' )
          #sys.exit( 1 )



      try:
        resp = service.removeUser(userid = userid)
        print('\nremoveUser response:')
        print(resp, '\n')
      except Fault as err:
        print( f'Zeep error: removeUser: { err }' )
        #sys.exit( 1 )


      for item in phone_prefixes:
         device_name = str(item + userid)
         try:
            resp = service.removePhone(name = device_name)
            print('\nremovePhone response:')
            print(resp, '\n')
         except Fault as err:
            print( f'Zeep error: removePhone: { err }' )
            #sys.exit( 1 )




def deleteConfig(provisioning_file, roomId):

    #initialize cucm parameters
    disable_warnings(InsecureRequestWarning)
    partition = credentials["partition"]

    wb_obj = openpyxl.load_workbook(provisioning_file)
    sheet = wb_obj.active

    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)

    print(col_names)
    width = len (col_names)
    if width >= 4:
       userid_flag = True
    data = {}

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            for j in range(width):
               data[row[j]] = []

        else:
            for j in range(width):
               data[col_names[j]].append(row[j])

    print (data)

    n = len (data['First Name'])

    print(n)

    for o in range(n):
       first_name = data['First Name'][o]
       last_name = data['Last Name'][o]
       tel = data['Phone Number'][o]
       if "Partition" in data.keys():
           if data['Partition'][o] != None:
               user_partition = data['Partition'][o]
           else:
               user_partition = partition
       else:
           user_partition = partition
       devices = {
           'device': []
                 }

       if data['UserID'][o] == None:
           l = len(last_name)
           print('lunghezza del cognome: ', l)
           if l < 8:
               userid = first_name[0].lower() + last_name.lower()
           else:
               userid = first_name[0].lower() + last_name[0:7].lower()
       else:
           userid = data['UserID'][o]

       email = userid + '@cloudentpa.com'
       delete_ldap_user(first_name, last_name, userid)
       deleteInbox(userid)
       if data['Provisioning'][o] != 'DI':
          #remove WxC Professional licenses
          deleteLicenses(email, roomId)
       if data['Provisioning'][o] == 'DI':
          #remove DI config
          deleteDI(userid, tel, user_partition)

username = credentials['username']
password = credentials['password']

# If you're not disabling SSL verification, host should be the FQDN of the server rather than IP
host = credentials['host']

#
org_url = credentials['org_url']
people_url = credentials['people_url']
user_details_url = credentials['user_details_url']
location_id_url = credentials['location_id_url']
license_url = credentials['license_url']



