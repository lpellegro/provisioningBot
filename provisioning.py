from zeep import Client
from zeep.cache import SqliteCache
from zeep.transports import Transport
from zeep.exceptions import Fault
from zeep.plugins import HistoryPlugin
import requests
from requests import Session
from requests.auth import HTTPBasicAuth
from urllib3 import disable_warnings
from urllib3.exceptions import InsecureRequestWarning
from lxml import etree
import boto3
import botocore
import openpyxl
from pathlib import Path
import json
from credentials import credentials
import get_tokens
from get_tokens import get_my_access_token
from useradd import connect_ldap_server
import time
from messages import messages, messageText, messageFile, webexHeaders, simpleHeaders

#provisioning the system
username = credentials['username']
password = credentials['password']
domain = credentials['domain']
host = credentials['host']
mailbox_provisioning = True
org_provisioning = True
ucm_provisioning = True
org_url = credentials['org_url']
people_url = credentials['people_url']
user_details_url = credentials['user_details_url']
location_id_url = credentials['location_id_url']
license_url = credentials['license_url']
webexUrl = credentials["webexUrl"]
botBearer = credentials["bearer"]
botHeaders = webexHeaders(botBearer, "application/json")
msg = messages()
main_location = credentials['main_location']
user_password = credentials["user_default_password"]

def aws_mailboxes (first_name, last_name, userid, email):
   #provision inboxes in Amazon WorkMail
   client = boto3.client('workmail', region_name=credentials["aws_region_name"], aws_access_key_id=credentials["aws_access_key_id"],
         aws_secret_access_key=credentials["aws_secret_access_key"])
   try:
      response = client.create_user(
        OrganizationId=credentials["aws_organization_id"],
        Name = userid,
        DisplayName = first_name + " " + last_name,
        Password=user_password
      )
      print('response is:', response)
      id_of_user = response['UserId']
      response = client.register_to_work_mail(
          OrganizationId=credentials["aws_organization_id"],
          EntityId=id_of_user,
          Email=email
      )
   except botocore.exceptions.ClientError as error:
       print (error, type(error))
       if "same" in str(error):
           pass

   except botocore.exceptions.ParamValidationError as error:
       raise ValueError('The parameters you provided are incorrect: {}'.format(error))


def webex_licenses (email, tel, headers, wxc_license, location):
   #Webex provisioning
   main_location = credentials["main_location"]
   if location != "multitenant":
      main_location = location
   response = requests.get(user_details_url + email, headers=headers)
   print(response.status_code)
   #if response.status_code == "404":
      #print(f"User {email} not found")
      #return
   response = response.json()
   print("user details", response)
   if response ["items"] == []:
     print(f"response to licenses for email {email} is {response}")
     return
   person_id = response['items'][0]['id']
   #provision licenses for new users in multitenant:
   person_licenses = response['items'][0]['licenses']
   #phoneNumber = response['items'][0]['phoneNumbers'][0]['value']
   #extension = response['items'][0]['extension']

   #if wxc_license not in person_licenses or phoneNumber[-4:] != extension:
   #if wxc_license not in person_licenses:
   user_data = response['items'][0]
   person_licenses.append(wxc_license)
   if "phoneNumbers" in user_data.keys():
        phoneNumber = response['items'][0]['phoneNumbers'][0]['value']
        extension = phoneNumber[-4:]
   else:
        extension = tel[-4:]
   del user_data['phoneNumbers']
   user_data['licenses'] = person_licenses
   user_data['extension'] = extension
   if 'displayName' not in user_data:
      user_data['displayName'] = ' '
   location_resp = requests.get(location_id_url, headers=headers)
   location_resp = location_resp.json()
   for i in range(len(location_resp['items'])):
       if location_resp['items'][i]['name'] == main_location:
          location_id = location_resp['items'][i]['id']
   user_data['locationId'] = location_id
   content_type_headers = headers
   content_type_headers['Content-Type'] = 'application/json'
   payload = json.dumps(user_data)
   print('payload is: ', payload)
   url = people_url + user_data['id'] + '?callingData=true'
   print('sending request to: ', url)
   response = requests.put(url, data=payload, headers=content_type_headers)
   print(response)
   print('put response is: ', response.json())


def provision_ucm(username, password, tel, first_name, last_name, userid, email, devices, existing_devices, partition, css, device_pool, ucm_location, sip_profile):
    print("Provisioning on DI:", username, password, tel, first_name, last_name, userid, email, devices, existing_devices, partition, css, device_pool, ucm_location, sip_profile)
    disable_warnings(InsecureRequestWarning)
    phone_prefixes = ['CSF', 'TAB', 'TCT', 'BOT']
    phones = {'CSF': 'Cisco Unified Client Services Framework', 'TAB': 'Cisco Jabber for Tablet',
              'TCT': 'Cisco Dual Mode for iPhone', 'BOT': 'Cisco Dual Mode for Android'}
    # device_name = str('CSF'+userid)
    wsdl = credentials['wsdl']

    location = 'https://{host}:8443/axl/'.format(host=host)

    binding = "{http://www.cisco.com/AXLAPIService/}AXLAPIBinding"
    if tel.startswith ("+"):
       normalized_tel = f"\{tel}"
       selfServiceId = tel.strip("+")
    else:
       normalized_tel = tel
       selfServiceId = tel

    # Create a custom session to disable Certificate verification.
    # In production you shouldn't do this,
    # but for testing it saves having to have the certificate in the trusted store.
    session = Session()
    session.verify = False
    session.auth = HTTPBasicAuth(username, password)

    transport = Transport(cache=SqliteCache(), session=session, timeout=20)
    history = HistoryPlugin()
    client = Client(wsdl=wsdl, transport=transport, plugins=[history])
    service = client.create_service(binding, location)

    def show_history():
        for item in [history.last_sent, history.last_received]:
            print(etree.tostring(item["envelope"], encoding="unicode", pretty_print=True))
    usercreateflag = False
    usermodifyflag = False
    linemodifyflag = False
    usercreate = ""
    try:
        resp = service.listUser(
            searchCriteria={
                'userid': userid
            },
            returnedTags={
                'userid': True,
                'telephoneNumber': True,
                'mailid': True,
                'directoryUri': True
            }

        )
        print(resp, '\n')
        if resp["return"] == None:
            print("User doesn't exists in directory")
            usercreate = "y" #input("Do you want to create a local user? ")
            if usercreate.lower().startswith("y"):
               usercreateflag = True
        else:
            telephone = resp["return"]["user"][0]["telephoneNumber"]
            #if telephone.startswith ("+"):
               #tel = f"\{telephone}"
            print("telephone in user config: ", telephone)
            print("telephone set on excel: ", tel)
            if telephone == None:
               usermodifyflag = True
               print(f"user {userid} lacks phone no")
            if telephone != tel:
               print("line needs to be updated")
               linemodifyflag = True
            directoryUri = resp["return"]["user"][0]["directoryUri"]
            email = resp["return"]["user"][0]["mailid"]
            print(f"telephone number is {tel} and email is {email}")
    except Exception as err:
        print(f'\nZeep error: listUser: {err}')
        #sys.exit(1)
        print('\nlistUser response:\n')

    if linemodifyflag == True:
        try:
            resp = service.updateLine(pattern = telephone, newPattern = normalized_tel)
            usermodifyflag = True

        except Fault as err:
            print(f'Zeep error: updateLine: {err}')
            #sys.exit(1)

        print('\nupdateLine response:\n')
        print(resp, '\n')

    else:

        try:
            resp = service.addLine(line={'pattern': normalized_tel,
                                         'usage': 'Device',
                                         'patternUrgency': True,
                                         'routePartitionName': partition,
                                         'shareLineAppearanceCssName': css})
            print("line creation response",resp)
            line_id = resp['return']
        except Fault as err:
            show_history()
            if "duplicate" in err.message:
                print ("duplicate in err")
                pass
    if usercreateflag == True:
        user = {
            'firstName': first_name,
            'lastName': last_name,
            'userid': userid,
            'password': user_password,
            'mailid': email,
            'telephoneNumber': tel,
            'directoryUri': email,
            'presenceGroupName': 'Standard Presence group',
            'associatedGroups': {
                'userGroup': [
                    {
                        'name': 'Customer CCM End Users',
                        'userRoles': {
                            'userRole': [
                                'Standard CCM Admin Users',
                                'Standard CCM End Users'
                            ]
                        }
                    }
                ],
                'userGroup': [
                    {
                        'name': 'Customer CCM Admin Users',
                        'userRoles': {
                            'userRole': [
                                'Customer AXL API Access',
                                'Customer CCDMADMIN',
                                'Customer IMPADMIN',
                                'Standard CCM Admin Users',
                                'Standard CCM End Users',
                                'Standard CCMUSER Administration',
                                'Standard CTI Allow Control of Phones supporting Connected Xfer and conf',
                                'Standard CTI Allow Control of Phones supporting Rollover Mode',
                                'Standard CTI Enabled',
                                'Standard CUReporting',
                                'Standard SERVICEABILITY Read Only'
                            ]
                        }
                    }
                ]
            }
        }
        usermodifyflag = True
        try:
            resp = service.addUser(user)

        except Fault as err:
            print(f'Zeep error: addUser: {err}')
            print('err.message is:', err.message)
            if "duplicate" in err.message:
                print ("duplicate user")
                pass
            #sys.exit(1)

    for item in phone_prefixes:
        device_name = str(item + userid.upper())
        device_type = phones[item]
        phone = {
            'name': device_name,
            'ownerUserName': userid,
            'product': device_type,
            'class': 'Phone',
            'protocol': 'SIP',
            'protocolSide': 'User',
            'devicePoolName': device_pool,
            'sipProfileName': sip_profile,
            'commonPhoneConfigName': 'Standard Common Phone Profile',
            'locationName': ucm_location,
            'useTrustedRelayPoint': 'Default',
            'builtInBridgeStatus': 'Default',
            'packetCaptureMode': 'None',
            'certificateOperation': 'No Pending Operation',
            'deviceMobilityMode': 'Default',
            'lines': {
                'line': {
                    'index': 1,
                    'dirn': {
                        'pattern': normalized_tel,
                        'routePartitionName': partition
                    },
                    'associatedEndusers': {
                        'enduser': {
                            'userId': userid
                        }
                    }
                }
            }
        }
        try:
            resp = service.addPhone(phone)
            devices['device'].append(device_name)
            print('devices are: ', devices['device'])
        except Fault as err:
            print(f'Zeep error: addPhone: {err}')
            if "duplicate" in err.message:
                print("duplicate phone")
                existing_devices.append(device_name)
                pass
            #sys.exit(1)

    devices["device"] = devices["device"] + existing_devices
    print(f"device list to be added to {userid}: {devices}")
#.... until here


    if usermodifyflag == True or linemodifyflag == True:
      try:
        resp = service.updateUser(
            userid=userid,
            associatedDevices=devices,
            homeCluster=True,
            imAndPresenceEnable=False,
            telephoneNumber = tel,
            primaryExtension={
                'pattern': normalized_tel,
                'routePartitionName': partition
            }
        )
        print(resp, '\n')

      except Exception as err:
        print(f'\nZeep error: updateUser: {err}')
        #sys.exit(1)
        print('\nupdateUser response:\n')

        if "duplicate" in err.message:
            print ("duplicate in err")
            pass

      try:
          resp = service.updateUser(
              userid=userid,
              associatedDevices=devices,
              homeCluster=True,
              imAndPresenceEnable=False,
              telephoneNumber=tel,
              selfService=selfServiceId,
              primaryExtension={
                  'pattern': normalized_tel,
                  'routePartitionName': partition
              }
          )
          print(resp, '\n')

      except Exception as err:
          print(f'\nZeep error: updateUser: {err}')
          # sys.exit(1)
          print('\nupdateUser response:\n')

          if "duplicate" in err.message:
              print("duplicate in err")
              pass


def provisioning(file, roomId):
    mailbox_provisioning = True
    org_provisioning = True
    ucm_provisioning = True
    ad_provisioning = True
    partition = credentials['partition']
    css = credentials['css']
    device_pool = credentials['device_pool']
    ucm_location = credentials['location']
    sip_profile = credentials['sip_profile']

    if mailbox_provisioning == False and ucm_provisioning == False and org_provisioning == False and ad_provisioning == False:
       exit()

    # If you're not disabling SSL verification, host should be the FQDN of the server rather than IP
    if org_provisioning == True:
        bearer = get_my_access_token(roomId)
        print("auth bearer is", bearer)
        headers = {'Authorization': 'Bearer ' + bearer}


        response = requests.get(org_url, headers=headers)
        response = response.json()
        print(response)
        #org_id = response['items'][0]['id']

        response = requests.get(license_url, headers=headers)
        response = response.json()
        licenses_list = response['items']
        l = len(licenses_list)
        for i in range (l):
            if licenses_list[i]['name'] == "Webex Calling - Professional":
                wxc_license = licenses_list[i]['id']
                break

    wb_obj = openpyxl.load_workbook(file)
    sheet = wb_obj.active

    col_names = []
    for column in sheet.iter_cols(1, sheet.max_column):
        col_names.append(column[0].value)

    print(col_names)
    width = len (col_names)
    if width >= 4:
       userid_flag = True
    data = {}

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            for j in range(width):
               data[row[j]] = []

        else:
            for j in range(width):
               data[col_names[j]].append(row[j])

    print ("USER DATA:",data)
    
    n = len (data['First Name'])

    print("n of items in First Name", n)
    emails = []
    for i in range(n):
       first_name = data['First Name'][i]
       last_name = data['Last Name'][i]
       tel = data['Phone Number'][i]

       if data['UserID'][i] == None:
           if last_name == None:
               break
           l = len(last_name)

           print('lunghezza del cognome: ', l)
           if l < 8:
               userid = first_name[0].lower() + last_name.lower()
           else:
               userid = first_name[0].lower() + last_name[0:7].lower()
       else:
           userid = data['UserID'][i]
       print(userid)
       email = userid + '@' + domain
       emails.append(email)
       if ad_provisioning == True and i == 0:
          print("PROVISIONING AD")
          msgText = messageText(roomId, msg.adProvisioningStart)
          content = msgText.data
          body = json.dumps(content)
          response = requests.post(f"{webexUrl}messages", headers=botHeaders.webexHeaders, data=body)
          print(response.text)
       if ad_provisioning == True:
          connect_ldap_server(first_name, last_name, userid, email, tel)
    data["email"] = emails
    msgText = messageText(roomId, msg.doneWait)
    content = msgText.data
    body = json.dumps(content)
    response = requests.post(f"{webexUrl}messages", headers=botHeaders.webexHeaders, data=body)
    print(response.text)
    print("waiting 6 minutes for DirSync")
    time.sleep(360)
    dimessage = False
    awsmessage = False
    wxcmessage = False
    for i in range(n):
       first_name = data['First Name'][i]
       last_name = data['Last Name'][i]
       tel = data['Phone Number'][i]
       userid = data['UserID'][i]
       email = data['email'][i]
       if data['Provisioning'][i] == 'DI':
           if "Partition" in data.keys():
               if data['Partition'][i] != None:
                   user_partition = data['Partition'][i]
               else:
                   user_partition = partition
           else:
               user_partition = partition
           if "Calling Search Space" in data.keys():
               if data['Calling Search Space'][i] != None:
                   user_css = data['Calling Search Space'][i]
               else:
                   user_css = css
           else:
               user_css = css
           if "Device Pool" in data.keys():
               if data['Device Pool'][i] != None:
                   user_device_pool = data['Device Pool'][i]
               else:
                   user_device_pool = device_pool
           else:
               user_device_pool = device_pool
           if "UCM Location" in data.keys():
               if data['UCM Location'][i] != None:
                   user_ucm_location = data['UCM Location'][i]
               else:
                   user_ucm_location = ucm_location
           else:
               user_ucm_location = ucm_location
           if "SIP Profile" in data.keys():
               if data['SIP Profile'][i] != None:
                   user_sip_profile = data['SIP Profile'][i]
               else:
                   user_sip_profile = sip_profile
           else:
               user_sip_profile = sip_profile

       devices = {
           'device': []
                 }
       print(f"Provisioning {first_name}, {last_name}, {userid}, {email}")

       existing_devices = []
       if mailbox_provisioning == True and awsmessage == False:
          print("PROVISIONING MAILBOXES")
          msgText = messageText(roomId, msg.awsProvisioningStart)
          content = msgText.data
          body = json.dumps(content)
          response = requests.post(f"{webexUrl}messages", headers=botHeaders.webexHeaders, data=body)
          print(response.text)
          awsmessage = True
       if mailbox_provisioning == True:
          print("PROVISIONING MAILBOXES")
          aws_mailboxes (first_name, last_name, userid, email)

       if data['Provisioning'][i] != 'DI' and org_provisioning == True and wxcmessage == False:
          print("PROVISIONING MULTITENANT")
          msgText = messageText(roomId, msg.orgProvisioningStart)
          content = msgText.data
          body = json.dumps(content)
          response = requests.post(f"{webexUrl}messages", headers=botHeaders.webexHeaders, data=body)
          print(response.text)
          wxcmessage = True
       if data['Provisioning'][i] != 'DI' and org_provisioning == True:
          location = data['Provisioning'][i]
          print(f"PROVISIONING MULTITENANT: {email}, {location}")
          webex_licenses (email, tel, headers, wxc_license, location)

       if data['Provisioning'][i] == 'DI' and ucm_provisioning == True and dimessage == False:
          print("PROVISIONING DEDICATED INSTANCE")
          msgText = messageText(roomId, msg.diProvisioningStart)
          content = msgText.data
          body = json.dumps(content)
          response = requests.post(f"{webexUrl}messages", headers=botHeaders.webexHeaders, data=body)
          print(response.text)
          dimessage = True
       if data['Provisioning'][i] == 'DI' and ucm_provisioning == True:
          print("PROVISIONING DEDICATED INSTANCE")
          provision_ucm (username, password, tel, first_name, last_name, userid, email, devices, existing_devices, user_partition, user_css, user_device_pool, user_ucm_location, user_sip_profile)





